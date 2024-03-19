# #!/usr/bin/env python3
# # -*- coding: utf-8 -*-

#    TODO:   对excel的读写操作
#    Author: LTH
#    Date:   2023-6-19


import xlrd  # 读取excel的xls格式
import openpyxl  # 读取excel的xlsx格式

from framerwork.log.log import Log


class Excel:
    # 初始化
    def __init__(self):
        self.path = None
        self.data = None

    # 创建excel
    @staticmethod
    def create(path, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(len(data)):
            for j in range(len(data[i])):
                ws.cell(row=i + 1, column=j + 1, value=str(data[i][j]))
        wb.save(path)

    # 打开excel
    def open(self, path, sheet=0):
        try:
            self.path = path
            if path.endswith(".xls"):
                return self.openXls(sheet)
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.worksheets[sheet]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            self.data = data
            return self.read()
        except Exception as e:
            Log.error("打开excel失败，原因：%s" % e)
            return None

    # 打开xls格式的excel
    def openXls(self, sheet=0):
        workbook = xlrd.open_workbook(self.path)
        sheet = workbook.sheet_by_index(sheet)
        data = []
        for i in range(sheet.nrows):
            data.append(sheet.row_values(i))
        self.data = data
        return self.read()

    # 读取excel
    def read(self):
        return self.data




    # 写入excel
    # def write(self, data):
    #     if self.path.endswith(".xls"):
    #         self.writeXls(data)
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     for i in range(len(data)):
    #         for j in range(len(data[i])):
    #             ws.cell(row=i + 1, column=j + 1, value=str(data[i][j]))
    #     wb.save(self.path)
    #
    # # 写入xls格式的excel
    # def writeXls(self, data):
    #     wb = xlwt.Workbook()
    #     ws = wb.add_sheet('Sheet1')
    #     for i in range(len(data)):
    #         for j in range(len(data[i])):
    #             ws.write(i, j, data[i][j])
    #     wb.save(self.path)
